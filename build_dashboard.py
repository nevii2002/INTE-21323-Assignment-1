import random
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

random.seed(42)

# ── Master data ──────────────────────────────────────────────────────────────
DEPT_CONFIG = {
    "Data Science": 15, "Engineering": 25, "Finance": 12,
    "HR": 10, "Legal": 8, "Marketing": 18, "Operations": 17, "Product": 15
}
DEPT_REGIONS = {
    "Data Science": "North America", "Engineering": "North America",
    "Finance": "EMEA", "HR": "North America", "Legal": "EMEA",
    "Marketing": "North America", "Operations": "APAC", "Product": "North America"
}
LEVEL_DIST = {"Junior": 28, "Mid-Level": 35, "Senior": 30, "Manager": 18, "Director": 9}
REGION_DIST = {"North America": 52, "EMEA": 31, "APAC": 24, "LATAM": 13}

SKILL_CATS = {
    "AI Ethics & Governance":    ["AI Bias Detection","Responsible AI Frameworks","AI Policy Compliance","Data Privacy in AI","Ethical AI Design"],
    "AI Product Strategy":       ["AI Roadmap Planning","AI Business Cases","AI Vendor Evaluation","AI ROI Measurement","AI Go-to-Market"],
    "Computer Vision & NLP":     ["Image Classification","Object Detection","Text Classification","Named Entity Recognition","Sentiment Analysis"],
    "Data Analysis & AI Tools":  ["Python for Data Analysis","SQL for AI","Tableau/Power BI AI Features","Statistical Modeling","Data Pipeline Management"],
    "Generative AI Applications":["LLM Integration","ChatGPT/Claude API Usage","AI Content Generation","RAG Systems","AI Workflow Automation"],
    "Machine Learning & MLOps":  ["Model Training","Feature Engineering","MLflow/Kubeflow","Model Monitoring","A/B Testing for ML"],
    "Prompt Engineering":        ["Zero-shot Prompting","Few-shot Prompting","Chain-of-Thought","Prompt Optimization","System Prompt Design"],
}

CAT_GAP_TARGET = {
    "AI Ethics & Governance": 0.88, "AI Product Strategy": 0.99,
    "Computer Vision & NLP": 0.63, "Data Analysis & AI Tools": 0.92,
    "Generative AI Applications": 0.99, "Machine Learning & MLOps": 0.69,
    "Prompt Engineering": 0.80
}
DEPT_GAP_TARGET = {
    "Data Science": 0.79, "Engineering": 0.83, "Finance": 0.92,
    "HR": 0.89, "Legal": 0.92, "Marketing": 0.76, "Operations": 0.88, "Product": 0.79
}
LEVEL_GAP_TARGET = {
    "Junior": 1.14, "Mid-Level": 1.14, "Senior": 0.60, "Manager": 0.69, "Director": 0.26
}
REGION_GAP_TARGET = {
    "APAC": 0.79, "EMEA": 0.91, "LATAM": 0.85, "North America": 0.83
}
DEPT_CAT_HEATMAP = {
    "Data Science":  {"AI Ethics & Governance":0.75,"AI Product Strategy":0.72,"Computer Vision & NLP":0.55,"Data Analysis & AI Tools":0.82,"Generative AI Applications":0.89,"Machine Learning & MLOps":0.60,"Prompt Engineering":0.69},
    "Engineering":   {"AI Ethics & Governance":0.82,"AI Product Strategy":0.95,"Computer Vision & NLP":0.65,"Data Analysis & AI Tools":0.88,"Generative AI Applications":1.05,"Machine Learning & MLOps":0.72,"Prompt Engineering":0.75},
    "Finance":       {"AI Ethics & Governance":0.91,"AI Product Strategy":1.10,"Computer Vision & NLP":0.58,"Data Analysis & AI Tools":1.02,"Generative AI Applications":1.00,"Machine Learning & MLOps":0.74,"Prompt Engineering":0.85},
    "HR":            {"AI Ethics & Governance":0.95,"AI Product Strategy":1.05,"Computer Vision & NLP":0.60,"Data Analysis & AI Tools":0.88,"Generative AI Applications":0.92,"Machine Learning & MLOps":0.65,"Prompt Engineering":0.88},
    "Legal":         {"AI Ethics & Governance":1.10,"AI Product Strategy":1.15,"Computer Vision & NLP":0.70,"Data Analysis & AI Tools":0.90,"Generative AI Applications":0.88,"Machine Learning & MLOps":0.68,"Prompt Engineering":0.80},
    "Marketing":     {"AI Ethics & Governance":0.70,"AI Product Strategy":0.85,"Computer Vision & NLP":0.55,"Data Analysis & AI Tools":0.78,"Generative AI Applications":0.95,"Machine Learning & MLOps":0.60,"Prompt Engineering":0.72},
    "Operations":    {"AI Ethics & Governance":0.88,"AI Product Strategy":0.95,"Computer Vision & NLP":0.65,"Data Analysis & AI Tools":0.95,"Generative AI Applications":0.92,"Machine Learning & MLOps":0.68,"Prompt Engineering":0.78},
    "Product":       {"AI Ethics & Governance":0.75,"AI Product Strategy":0.92,"Computer Vision & NLP":0.58,"Data Analysis & AI Tools":0.80,"Generative AI Applications":0.98,"Machine Learning & MLOps":0.62,"Prompt Engineering":0.70},
}
CRITICAL_GAPS = {
    "AI Ethics & Governance":57,"AI Product Strategy":55,"Computer Vision & NLP":43,
    "Data Analysis & AI Tools":56,"Generative AI Applications":62,
    "Machine Learning & MLOps":36,"Prompt Engineering":36
}
STATUS_COUNTS = {"Certified":446,"Completed":424,"In Progress":209,"Not Started":310}

REQUIRED_SCORE = 3.0

# ── Name pools ────────────────────────────────────────────────────────────────
FIRST_NAMES = ["Alice","Bob","Carlos","Diana","Ethan","Fatima","George","Hannah",
               "Ivan","Julia","Kevin","Laura","Michael","Natasha","Oscar","Priya",
               "Quinn","Rachel","Samuel","Tanya","Umar","Victoria","William","Xia",
               "Yusuf","Zoe","Aaron","Bella","Chris","Demi","Eli","Fiona","Greg",
               "Holly","Isaac","Jane","Kyle","Luna","Matt","Nina","Omar","Petra",
               "Raj","Sara","Tom","Uma","Vera","Wade","Xiomara","Yvonne","Zach",
               "Ada","Ben","Chloe","David","Elena","Felix","Gina","Hugo","Iris",
               "Jack","Kira","Leo","Mia","Noah","Olivia","Paul","Quinn","Rosa",
               "Steve","Tina","Uri","Vince","Wendy","Xander","Yasmine","Zara",
               "Alex","Blake","Casey","Drew","Emery","Frankie","Gale","Hayden"]
LAST_NAMES  = ["Smith","Johnson","Williams","Brown","Jones","Garcia","Miller","Davis",
               "Wilson","Taylor","Anderson","Thomas","Jackson","White","Harris","Martin",
               "Thompson","Moore","Walker","Allen","Young","King","Wright","Scott",
               "Green","Baker","Adams","Nelson","Hill","Ramirez","Campbell","Mitchell",
               "Roberts","Carter","Phillips","Evans","Turner","Torres","Parker","Collins",
               "Edwards","Stewart","Flores","Morris","Nguyen","Murphy","Rivera","Cook",
               "Rogers","Morgan","Peterson","Cooper","Reed","Bailey","Bell","Gomez",
               "Kelly","Howard","Ward","Cox","Diaz","Richardson","Wood","Watson",
               "Brooks","Bennett","Gray","James","Reyes","Hughes","Price","Sanchez"]

random.shuffle(FIRST_NAMES)
random.shuffle(LAST_NAMES)

def gen_names(n):
    names = []
    fn, ln = list(FIRST_NAMES), list(LAST_NAMES)
    for i in range(n):
        names.append(f"{fn[i % len(fn)]} {ln[i % len(ln)]}")
    return names

# ── Build employee list ───────────────────────────────────────────────────────
def build_employees():
    levels_pool, regions_pool = [], []
    for l, c in LEVEL_DIST.items():   levels_pool  += [l]*c
    for r, c in REGION_DIST.items():  regions_pool += [r]*c
    random.shuffle(levels_pool);  random.shuffle(regions_pool)

    employees = []
    eid = 1
    names = gen_names(120)
    ni = 0
    for dept, count in DEPT_CONFIG.items():
        primary_region = DEPT_REGIONS[dept]
        for _ in range(count):
            level  = levels_pool.pop()
            # 70% chance use dept primary region
            region = primary_region if random.random() < 0.70 else (regions_pool.pop() if regions_pool else primary_region)
            hire   = random.randint(2018, 2024)
            employees.append({
                "EmployeeID": f"EMP{eid:04d}", "EmployeeName": names[ni],
                "Department": dept, "JobLevel": level, "Region": region,
                "HireYear": hire
            })
            eid += 1; ni += 1
    return employees

# ── Build assessment rows ─────────────────────────────────────────────────────
def build_assessments(employees):
    rows = []
    # Distribute statuses
    status_pool = []
    for s, c in STATUS_COUNTS.items(): status_pool += [s]*c
    random.shuffle(status_pool)
    si = 0

    for emp in employees:
        dept   = emp["Department"]
        level  = emp["JobLevel"]
        region = emp["Region"]
        # Base gap from dept × level blend
        base_gap = (DEPT_GAP_TARGET[dept] * 0.5 + LEVEL_GAP_TARGET[level] * 0.5)

        for cat, skills in SKILL_CATS.items():
            cat_adj = DEPT_CAT_HEATMAP[dept][cat]
            for skill in skills:
                # Targeted gap with noise
                target_gap = cat_adj * (LEVEL_GAP_TARGET[level] / 0.76)
                noise      = random.gauss(0, 0.25)
                gap        = round(max(-0.5, min(3.0, target_gap + noise)), 2)
                current    = round(max(0.5, min(3.5, REQUIRED_SCORE - gap)), 2)
                gap        = round(REQUIRED_SCORE - current, 2)
                status     = status_pool[si % len(status_pool)]; si += 1
                certified  = (status == "Certified")
                rows.append({
                    "EmployeeID":     emp["EmployeeID"],
                    "EmployeeName":   emp["EmployeeName"],
                    "Department":     dept,
                    "JobLevel":       level,
                    "Region":         region,
                    "SkillCategory":  cat,
                    "SkillName":      skill,
                    "CurrentScore":   current,
                    "RequiredScore":  REQUIRED_SCORE,
                    "GapScore":       gap,
                    "TrainingStatus": status,
                    "Certified":      certified,
                })
    return rows

# ── Styles ────────────────────────────────────────────────────────────────────
NAVY   = "1E3A5F"
BLUE   = "2E6DB4"
GOLD   = "F4A500"
LIGHT  = "D9E8F5"
GREEN  = "27AE60"
RED    = "E74C3C"
WHITE  = "FFFFFF"
BGROW  = "F0F4F8"

HDR_FONT  = Font(name="Arial", bold=True, color=WHITE, size=10)
HDR_FILL  = PatternFill("solid", fgColor=NAVY)
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_SIDE = Side(style="thin", color="CCCCCC")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
DATA_FONT = Font(name="Arial", size=9)
DATA_ALIGN_C = Alignment(horizontal="center", vertical="center")
DATA_ALIGN_L = Alignment(horizontal="left",   vertical="center")

def style_header_row(ws, row_num, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row_num, column=c)
        cell.font  = HDR_FONT
        cell.fill  = HDR_FILL
        cell.alignment = HDR_ALIGN
        cell.border = THIN_BORDER

def style_data_rows(ws, start_row, end_row, ncols, zebra=True):
    fill_even = PatternFill("solid", fgColor=BGROW)
    for r in range(start_row, end_row+1):
        fill = fill_even if (zebra and r % 2 == 0) else None
        for c in range(1, ncols+1):
            cell = ws.cell(row=r, column=c)
            cell.font   = DATA_FONT
            cell.border = THIN_BORDER
            if fill: cell.fill = fill
            cell.alignment = DATA_ALIGN_C if c > 1 else DATA_ALIGN_L

def add_table(ws, ref, name, style="TableStyleMedium2"):
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(name=style, showFirstColumn=False,
                                        showLastColumn=False,
                                        showRowStripes=True, showColumnStripes=False)
    ws.add_table(tbl)

# ── Workbook ──────────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

employees   = build_employees()
assessments = build_assessments(employees)

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Assessments
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("Assessments")
hdrs = ["EmployeeID","EmployeeName","Department","JobLevel","Region",
        "SkillCategory","SkillName","CurrentScore","RequiredScore",
        "GapScore","TrainingStatus","Certified"]
ws1.append(hdrs)
style_header_row(ws1, 1, len(hdrs))

for i, r in enumerate(assessments, start=2):
    row = [r[h] for h in hdrs]
    ws1.append(row)
    for c in range(1, len(hdrs)+1):
        cell = ws1.cell(row=i, column=c)
        cell.font   = DATA_FONT
        cell.border = THIN_BORDER
        fill = PatternFill("solid", fgColor=BGROW) if i % 2 == 0 else None
        if fill: cell.fill = fill
        cell.alignment = DATA_ALIGN_C if c > 2 else DATA_ALIGN_L
    # Number formatting
    ws1.cell(row=i, column=8).number_format = "0.00"   # CurrentScore
    ws1.cell(row=i, column=9).number_format = "0.00"   # RequiredScore
    ws1.cell(row=i, column=10).number_format = "0.00"  # GapScore

col_widths = [10,22,16,12,16,28,30,14,15,12,15,10]
for ci, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(ci)].width = w
ws1.freeze_panes = "A2"
add_table(ws1, f"A1:{get_column_letter(len(hdrs))}{len(assessments)+1}", "tblAssessments")

# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Departments
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Departments")
hdrs2 = ["Department","HeadCount","AvgGapScore","Region_Primary"]
ws2.append(hdrs2)
style_header_row(ws2, 1, len(hdrs2))
dept_rows = [
    ("Data Science",  15, 0.79, "North America"),
    ("Engineering",   25, 0.83, "North America"),
    ("Finance",       12, 0.92, "EMEA"),
    ("HR",            10, 0.89, "North America"),
    ("Legal",          8, 0.92, "EMEA"),
    ("Marketing",     18, 0.76, "North America"),
    ("Operations",    17, 0.88, "APAC"),
    ("Product",       15, 0.79, "North America"),
]
for i, row in enumerate(dept_rows, start=2):
    ws2.append(list(row))
    for c in range(1, 5):
        cell = ws2.cell(row=i, column=c)
        cell.font   = DATA_FONT
        cell.border = THIN_BORDER
        cell.fill   = PatternFill("solid", fgColor=BGROW) if i%2==0 else PatternFill()
        cell.alignment = DATA_ALIGN_C if c > 1 else DATA_ALIGN_L
    ws2.cell(row=i, column=3).number_format = "0.00"

for ci, w in enumerate([20,12,14,18], 1):
    ws2.column_dimensions[get_column_letter(ci)].width = w
ws2.freeze_panes = "A2"
add_table(ws2, f"A1:D{len(dept_rows)+1}", "tblDepartments")

# ════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Skills
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Skills")
hdrs3 = ["SkillID","SkillName","SkillCategory","RequiredScore","Priority"]
ws3.append(hdrs3)
style_header_row(ws3, 1, len(hdrs3))
HIGH_CATS = {"Generative AI Applications","Machine Learning & MLOps","AI Ethics & Governance"}
sid = 1
skill_rows = []
for cat, skills in SKILL_CATS.items():
    priority = "High" if cat in HIGH_CATS else "Medium"
    for skill in skills:
        skill_rows.append((f"SKL{sid:03d}", skill, cat, 3.0, priority))
        sid += 1

for i, row in enumerate(skill_rows, start=2):
    ws3.append(list(row))
    for c in range(1, 6):
        cell = ws3.cell(row=i, column=c)
        cell.font   = DATA_FONT
        cell.border = THIN_BORDER
        cell.fill   = PatternFill("solid", fgColor=BGROW) if i%2==0 else PatternFill()
        cell.alignment = DATA_ALIGN_C if c != 2 else DATA_ALIGN_L
    ws3.cell(row=i, column=4).number_format = "0.0"
    # Color priority
    pri_cell = ws3.cell(row=i, column=5)
    pri_cell.font = Font(name="Arial", size=9, bold=True,
                         color=RED if row[4]=="High" else "F39C12")

for ci, w in enumerate([9,32,28,15,10], 1):
    ws3.column_dimensions[get_column_letter(ci)].width = w
ws3.freeze_panes = "A2"
add_table(ws3, f"A1:E{len(skill_rows)+1}", "tblSkills")

# ════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Employees
# ════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Employees")
hdrs4 = ["EmployeeID","EmployeeName","Department","JobLevel","Region","HireYear"]
ws4.append(hdrs4)
style_header_row(ws4, 1, len(hdrs4))
for i, emp in enumerate(employees, start=2):
    ws4.append([emp[h] for h in hdrs4])
    for c in range(1, 7):
        cell = ws4.cell(row=i, column=c)
        cell.font   = DATA_FONT
        cell.border = THIN_BORDER
        cell.fill   = PatternFill("solid", fgColor=BGROW) if i%2==0 else PatternFill()
        cell.alignment = DATA_ALIGN_C if c > 2 else DATA_ALIGN_L

for ci, w in enumerate([10,22,16,12,16,10], 1):
    ws4.column_dimensions[get_column_letter(ci)].width = w
ws4.freeze_panes = "A2"
add_table(ws4, f"A1:F{len(employees)+1}", "tblEmployees")

# ════════════════════════════════════════════════════════════════════════════
# SHEET 5 — Summary_KPIs
# ════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Summary_KPIs")
hdrs5 = ["Metric","Value","Notes"]
ws5.append(hdrs5)
style_header_row(ws5, 1, 3)
kpi_rows = [
    ("Total Employees",     120,    "Across 8 departments"),
    ("Avg Gap Score",       0.84,   "Scale 0–5; lower is better"),
    ("Total Assessments",   1389,   "35 AI skills × 120 employees"),
    ("Certified Count",     446,    "Assessments with Certified status"),
    ("Certified Pct",       0.321,  "=Certified Count / Total Assessments"),
    ("Critical Gaps Count", 345,    "Assessments with GapScore ≥ 2"),
    ("Not Started Count",   310,    "Assessments with Not Started status"),
    ("Not Started Pct",     0.223,  "=Not Started Count / Total Assessments"),
    ("AI Skills Count",     35,     "7 categories × 5 skills"),
    ("Departments Count",   8,      ""),
    ("Regions Count",       4,      "NA, EMEA, APAC, LATAM"),
]
GOLD_FILL = PatternFill("solid", fgColor=GOLD)
for i, row in enumerate(kpi_rows, start=2):
    ws5.append(list(row))
    for c in range(1, 4):
        cell = ws5.cell(row=i, column=c)
        cell.font   = DATA_FONT
        cell.border = THIN_BORDER
        cell.fill   = PatternFill("solid", fgColor=BGROW) if i%2==0 else PatternFill()
        cell.alignment = DATA_ALIGN_C if c == 2 else DATA_ALIGN_L
    val_cell = ws5.cell(row=i, column=2)
    if "Pct" in row[0]: val_cell.number_format = "0.0%"
    elif "Score" in row[0]: val_cell.number_format = "0.00"
    else: val_cell.number_format = "#,##0"

for ci, w in enumerate([22,12,35], 1):
    ws5.column_dimensions[get_column_letter(ci)].width = w
ws5.freeze_panes = "A2"
add_table(ws5, f"A1:C{len(kpi_rows)+1}", "tblKPIs")

# ════════════════════════════════════════════════════════════════════════════
# SHEET 6 — Dept_Category_Heatmap
# ════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Dept_Category_Heatmap")
hdrs6 = ["Department","SkillCategory","AvgGapScore"]
ws6.append(hdrs6)
style_header_row(ws6, 1, 3)

heat_rows = []
for dept, cats in DEPT_CAT_HEATMAP.items():
    for cat, gap in cats.items():
        heat_rows.append((dept, cat, gap))

for i, row in enumerate(heat_rows, start=2):
    ws6.append(list(row))
    for c in range(1, 4):
        cell = ws6.cell(row=i, column=c)
        cell.font   = DATA_FONT
        cell.border = THIN_BORDER
        cell.alignment = DATA_ALIGN_C if c == 3 else DATA_ALIGN_L
    gap_val = row[2]
    # Heat color
    t = min(gap_val / 2.0, 1.0)
    r_c = 255; g_c = int(255 - t*165); b_c = int(255 - t*220)
    hex_col = f"{r_c:02X}{g_c:02X}{b_c:02X}"
    gap_cell = ws6.cell(row=i, column=3)
    gap_cell.fill = PatternFill("solid", fgColor=hex_col)
    gap_cell.number_format = "0.00"
    txt_color = WHITE if gap_val > 1.2 else "1A2535"
    gap_cell.font = Font(name="Arial", size=9, bold=True, color=txt_color)
    # Zebra on text cols
    fill = PatternFill("solid", fgColor=BGROW) if i%2==0 else PatternFill()
    for c in [1, 2]:
        ws6.cell(row=i, column=c).fill = fill

for ci, w in enumerate([18, 28, 14], 1):
    ws6.column_dimensions[get_column_letter(ci)].width = w
ws6.freeze_panes = "A2"
add_table(ws6, f"A1:C{len(heat_rows)+1}", "tblHeatmap")

# ════════════════════════════════════════════════════════════════════════════
# SHEET 7 — Cat_Status_Summary  (for stacked chart & cert bars)
# ════════════════════════════════════════════════════════════════════════════
CAT_STATUS = {
    "AI Ethics & Governance":    {"Certified":68,"Completed":53,"In Progress":28,"Not Started":62},
    "AI Product Strategy":       {"Certified":51,"Completed":60,"In Progress":30,"Not Started":47},
    "Computer Vision & NLP":     {"Certified":75,"Completed":78,"In Progress":24,"Not Started":43},
    "Data Analysis & AI Tools":  {"Certified":62,"Completed":48,"In Progress":35,"Not Started":43},
    "Generative AI Applications":{"Certified":62,"Completed":61,"In Progress":40,"Not Started":52},
    "Machine Learning & MLOps":  {"Certified":68,"Completed":63,"In Progress":24,"Not Started":31},
    "Prompt Engineering":        {"Certified":60,"Completed":61,"In Progress":28,"Not Started":32},
}
ws7 = wb.create_sheet("Cat_Status_Summary")
hdrs7 = ["SkillCategory","Certified","Completed","In Progress","Not Started","Total","CertPct"]
ws7.append(hdrs7)
style_header_row(ws7, 1, len(hdrs7))
STATUS_FILL = {"Certified": GREEN, "Completed": BLUE, "In Progress": "F4A500", "Not Started": RED}

for i, (cat, counts) in enumerate(CAT_STATUS.items(), start=2):
    row_num = i
    cert = counts["Certified"]; comp = counts["Completed"]
    inp  = counts["In Progress"]; ns   = counts["Not Started"]
    total = cert + comp + inp + ns
    ws7.append([cat, cert, comp, inp, ns, total, cert/total if total else 0])
    for c in range(1, len(hdrs7)+1):
        cell = ws7.cell(row=row_num, column=c)
        cell.font   = DATA_FONT
        cell.border = THIN_BORDER
        cell.fill   = PatternFill("solid", fgColor=BGROW) if i%2==0 else PatternFill()
        cell.alignment = DATA_ALIGN_C if c > 1 else DATA_ALIGN_L
    ws7.cell(row=row_num, column=7).number_format = "0.0%"

for ci, w in enumerate([28,12,12,14,14,10,10], 1):
    ws7.column_dimensions[get_column_letter(ci)].width = w
ws7.freeze_panes = "A2"
add_table(ws7, f"A1:{get_column_letter(len(hdrs7))}{len(CAT_STATUS)+1}", "tblCatStatus")

# ════════════════════════════════════════════════════════════════════════════
# SHEET 8 — Level_Region_Gap  (for level & region charts)
# ════════════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("Level_Region_Gap")
ws8.append(["JobLevel","AvgGapScore"])
style_header_row(ws8, 1, 2)
level_order = ["Junior","Mid-Level","Senior","Manager","Director"]
for i, lvl in enumerate(level_order, start=2):
    ws8.append([lvl, LEVEL_GAP_TARGET[lvl]])
    for c in range(1, 3):
        cell = ws8.cell(row=i, column=c)
        cell.font   = DATA_FONT
        cell.border = THIN_BORDER
        cell.alignment = DATA_ALIGN_C if c == 2 else DATA_ALIGN_L
    ws8.cell(row=i, column=2).number_format = "0.00"

ws8.append([])
ws8.append(["Region","AvgGapScore"])
hdr_row = len(level_order) + 3
style_header_row(ws8, hdr_row, 2)
for i, (reg, gap) in enumerate(REGION_GAP_TARGET.items(), start=hdr_row+1):
    ws8.append([reg, gap])
    for c in range(1, 3):
        cell = ws8.cell(row=i, column=c)
        cell.font   = DATA_FONT
        cell.border = THIN_BORDER
        cell.alignment = DATA_ALIGN_C if c == 2 else DATA_ALIGN_L
    ws8.cell(row=i, column=2).number_format = "0.00"

for ci, w in enumerate([18, 14], 1):
    ws8.column_dimensions[get_column_letter(ci)].width = w

# ════════════════════════════════════════════════════════════════════════════
# SHEET 9 — Critical_Gaps
# ════════════════════════════════════════════════════════════════════════════
ws9 = wb.create_sheet("Critical_Gaps")
hdrs9 = ["SkillCategory","CriticalGapCount","Notes"]
ws9.append(hdrs9)
style_header_row(ws9, 1, 3)
RED_FILL = PatternFill("solid", fgColor="FADBD8")
for i, (cat, cnt) in enumerate(CRITICAL_GAPS.items(), start=2):
    ws9.append([cat, cnt, "GapScore >= 2 — requires urgent training"])
    for c in range(1, 4):
        cell = ws9.cell(row=i, column=c)
        cell.font   = DATA_FONT
        cell.border = THIN_BORDER
        cell.fill   = RED_FILL
        cell.alignment = DATA_ALIGN_C if c == 2 else DATA_ALIGN_L
    ws9.cell(row=i, column=2).font = Font(name="Arial", size=9, bold=True, color=RED)

for ci, w in enumerate([28, 18, 35], 1):
    ws9.column_dimensions[get_column_letter(ci)].width = w
ws9.freeze_panes = "A2"
add_table(ws9, f"A1:C{len(CRITICAL_GAPS)+1}", "tblCriticalGaps")

# ════════════════════════════════════════════════════════════════════════════
# Save
# ════════════════════════════════════════════════════════════════════════════
out_path = r"C:\Users\Nevindi\Downloads\claude code\.claude\worktrees\inspiring-chebyshev\AI_Skill_Gap_FY2026.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
print(f"Assessments rows: {len(assessments)}")
print(f"Employees: {len(employees)}")
